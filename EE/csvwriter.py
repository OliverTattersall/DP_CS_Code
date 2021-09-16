import csv
import xlsxwriter
# import openpyxl

data = [
    ["SN", "Name", "Contribution"],
    [1, "Linus Torvalds", "Linux Kernel"],
    [2, "Tim Berners-Lee", "World Wide Web"],
    [3, "Guido van Rossum", "Python Programming"]
]


# with open('innovators.csv', 'w', newline='') as csvfile:
#     writr = csv.writer(csvfile, delimiter=' ', quotechar='"', quoting=csv.QUOTE_ALL)

#     writr.writerow(["SN", "Name", "Contribution"])
#     writr.writerow([1, "Linus Torvalds", "Linux Kernel"])
#     writr.writerow([2, "Tim Berners-Lee", "World Wide Web"])
#     writr.writerow([3, "Guido van Rossum", "Python Programming"])
# csvfile.close()


# workbook = xlsxwriter.Workbook('innovators.xlsx')
# worksheet2 = workbook.add_worksheet("Data")



# for i in range(len(data)):
#     for j in range(len(data[i])):
#         worksheet2.write(i, j, data[i][j])




# workbook.close()




from openpyxl import load_workbook
wb=load_workbook('innovators.xlsx')

sheet = wb['Data']
sheet['A8'] = ''
sheet['B20']=30


ws2 = wb["Fruits"]

fruits=["Banana", "Mango","Apple", "Pear"]
for i in range(4):
    ws2[chr(65+i)+str(i+1)]=fruits[i]


wb.save('innovators.xlsx')


