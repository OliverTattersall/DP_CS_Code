from openpyxl import load_workbook
from openpyxl import Workbook

import random
import time
import mergeSortcode
import quickSortcode



wb=Workbook()

row=1
ws= wb.active

# ### few duplicates


i=5000
while i<1000000:
    for _ in range(100):
        lst=[random.randint(1,1000000) for _ in range(i)]
        lst2=lst.copy()

        timeh = time.time()
        mergeSortcode.merge_sort(lst)
        timedh = time.time()-timeh

        timeq = time.time()
        quickSortcode.quicksort(lst2, 0, len(lst2)-1)
        timedq= time.time()-timeq

        tempdata=[i, timedh, timedq]


        for j in range(3):

            ws[chr(j+65)+str(row)]=tempdata[j]
        row+=1

    ws["B"+str(row)]="=SUM(B"+str(row-100)+":B"+str(row-1)+")/100"
    ws["C"+str(row)]="=SUM(C"+str(row-100)+":C"+str(row-1)+")/100"
    row+=2

    i=i*2

# workbook.close()

# wb=load_workbook("experimentData.xlsx")

sheet = wb.create_sheet("many duplicates")
### many duplicates
row=1
i=5000
while i<1000000:
    for _ in range(100):
        lst=[random.randint(1,i//100) for _ in range(i)]
        lst2=lst.copy()
        
        
        timeh = time.time()
        mergeSortcode.merge_sort(lst)
        timedh = time.time()-timeh

        timeq = time.time()
        quickSortcode.quicksort(lst2, 0, len(lst2)-1)
        timedq= time.time()-timeq

        tempdata=[i, timedh, timedq]
        for j in range(3):
            sheet[chr(65+j)+str(row)]=tempdata[j]
        row+=1
    sheet["B"+str(row)]="=SUM(B"+str(row-100)+":B"+str(row-1)+")/100"
    sheet["C"+str(row)]="=SUM(C"+str(row-100)+":C"+str(row-1)+")/100"
    row+=2
    i=i*2

# wb.save("experimentData.xlsx")
        




######  reversed and slightly changed

## reversed 
# wb=load_workbook("experimentData.xlsx")
sheet2 = wb.create_sheet("reversed and slight change")
i=5000
row = 1
while i < 1000000:
    lst =[n for n in range(i, 0, -1)]
    lst2=lst.copy()

    timeh = time.time()
    mergeSortcode.merge_sort(lst)
    timedh = time.time()-timeh

    timeq = time.time()
    quickSortcode.quicksort(lst2, 0, len(lst2)-1)
    timedq= time.time()-timeq


    tempdata=[i, timedh, timedq]
    for j in range(3):
        sheet2[chr(65+j)+str(row)]=tempdata[j]
    row+=1
    i=i*2


### slightly changed
i=5000
row=1
while i<1000000:
    lst = [n for n in range(i)]
    randoms = [random.randint(1,i-1) for _ in range(20)]
    for n in range(0, 10, 2):
        lst[randoms[n]], lst[randoms[n+1]] = lst[randoms[n+1]], lst[randoms[n]]
    
    lst2=lst.copy()
    timeh = time.time()
    mergeSortcode.merge_sort(lst)
    timedh = time.time()-timeh

    timeq = time.time()
    quickSortcode.quicksort(lst2, 0, len(lst2)-1)
    timedq= time.time()-timeq
    tempdata=[i, timedh, timedq]
    
    for j in range(3):
        sheet2[chr(72+j)+str(row)]=tempdata[j]
    row+=1
    i=i*2


wb.save("mergesortData.xlsx")